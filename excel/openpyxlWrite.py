from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
wb = Workbook()

# 获取当前活跃的worksheet,默认就是第一个worksheet
ws = wb.active

# 从第2行开始，写入9行10列数据，值为对应的列序号A、B、C、D...
for row in range(1, 10):
    for col in range(1, 11):
        ws.cell(row=row, column=col).value = get_column_letter(col)
# 保存
wb.save(filename="a.xls")
