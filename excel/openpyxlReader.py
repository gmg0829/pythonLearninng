from openpyxl import load_workbook

# 打开一个workbook
wb = load_workbook(filename="a.xlsx")

print(wb.sheetnames)